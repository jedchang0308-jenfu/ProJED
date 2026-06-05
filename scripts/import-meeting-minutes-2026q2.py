#!/usr/bin/env python
"""Import the 2026 Q2 management meeting minutes workbook into ProJED.

This is intentionally a one-off operational script. It avoids overwriting
existing imported rows by checking deterministic legacy_node_id values before
inserting anything.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_EMAIL = "jedchang0308@jenfu.com.tw"
DEFAULT_BOARD_NAME = "鉦富經營管理會議看板"
DEFAULT_BOARD_FALLBACK = "鉦富經營管理會議"
DEFAULT_ENV = ".env.p8.local"
DEFAULT_XLSX = Path("..") / "Meeting Minutes_經營管理會議_2026Q2.xlsx"
SHEET_NAME = "20260605"
IMPORT_PREFIX = "mm-2026q2-20260605"
SOURCE_LABEL = "meeting_minutes_2026q2"

STATUS_MAP = {
    "完成": ("completed", "完成"),
    "進行中": ("in_progress", "進行中"),
}

COLUMN_DEFS = {
    "待辦": {"legacy": f"{IMPORT_PREFIX}-column-todo", "status": "todo"},
    "進行中": {"legacy": f"{IMPORT_PREFIX}-column-in-progress", "status": "in_progress"},
    "完成": {"legacy": f"{IMPORT_PREFIX}-column-completed", "status": "completed"},
}


class ImportErrorWithContext(RuntimeError):
    pass


def load_env(path: Path) -> dict[str, str]:
    if not path.exists():
        raise ImportErrorWithContext(f"Env file not found: {path}")
    env: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        match = re.match(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.*)\s*$", line)
        if not match:
            continue
        value = match.group(2).strip()
        if (value.startswith('"') and value.endswith('"')) or (
            value.startswith("'") and value.endswith("'")
        ):
            value = value[1:-1]
        env[match.group(1)] = value
    return env


def json_safe(value: Any) -> Any:
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return value


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def compact_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_person(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "").lower()
    value = re.sub(r"\([^)]*\)", "", value)
    value = re.sub(r"（[^）]*）", "", value)
    value = re.sub(r"\[[^\]]*\]", "", value)
    value = re.sub(r"[^\w\u4e00-\u9fff@.-]+", "", value)
    return value.strip()


def person_tokens(raw: str) -> list[str]:
    raw = clean_text(raw)
    if not raw or raw.upper() in {"NA", "-"}:
        return []
    pieces = re.split(r"[\n,;/、]+", raw)
    tokens: list[str] = []
    for piece in pieces:
        normalized = normalize_person(piece)
        if normalized and normalized not in {"na", "-"}:
            tokens.append(normalized)
    return tokens


def member_match_index(members: list[dict[str, Any]]) -> dict[str, str]:
    index: dict[str, str] = {}
    for member in members:
        profile = member.get("profiles") or {}
        user_id = member["user_id"]
        email = clean_text(profile.get("email"))
        display = clean_text(profile.get("display_name"))
        candidates = {normalize_person(email), normalize_person(email.split("@")[0]), normalize_person(display)}
        for word in re.split(r"[\s\[\]（）()]+", unicodedata.normalize("NFKC", display).lower()):
            normalized = normalize_person(word)
            if normalized:
                candidates.add(normalized)
        for candidate in candidates:
            if candidate:
                index[candidate] = user_id
    return index


def resolve_people(raw: str, index: dict[str, str]) -> tuple[list[str], list[str]]:
    matched: list[str] = []
    unmatched: list[str] = []
    for token in person_tokens(raw):
        user_id = index.get(token)
        if user_id and user_id not in matched:
            matched.append(user_id)
        elif not user_id and token not in unmatched:
            unmatched.append(token)
    return matched, unmatched


def parse_target_date(value: Any) -> tuple[str | None, str]:
    if isinstance(value, dt.datetime):
        return value.date().isoformat(), value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat(), value.isoformat()
    return None, clean_text(value)


def parse_wbs_line(line: str) -> tuple[int, str]:
    stripped = line.strip()
    if not stripped:
        return 1, ""
    depth = 1
    title = stripped
    marker = re.match(r"^(\*+|-+)\s*(.*)$", stripped)
    if marker:
        depth = max(1, min(6, len(marker.group(1))))
        title = marker.group(2).strip()
    return depth, title or stripped


def parse_workbook(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise ImportErrorWithContext(f"Workbook not found: {path}")
    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ImportErrorWithContext(f"Sheet {SHEET_NAME!r} not found in {path}")
    ws = wb[SHEET_NAME]

    tasks: list[dict[str, Any]] = []
    skipped_rows: list[int] = []
    unknown_statuses: Counter[str] = Counter()
    status_raw_counts: Counter[str] = Counter()
    wbs_total = 0
    non_date_targets: list[dict[str, Any]] = []

    for row_num in range(3, ws.max_row + 1):
        title = clean_text(ws.cell(row_num, 1).value)
        status_raw = clean_text(ws.cell(row_num, 2).value)
        if not title and not status_raw:
            continue
        if not title or status_raw.upper() == "NA":
            skipped_rows.append(row_num)
            continue

        status_raw_counts[status_raw] += 1
        status, column = STATUS_MAP.get(status_raw, ("todo", "待辦"))
        if status_raw not in STATUS_MAP:
            unknown_statuses[status_raw] += 1

        end_date, target_time_raw = parse_target_date(ws.cell(row_num, 5).value)
        if target_time_raw and not end_date and target_time_raw.upper() != "NA":
            non_date_targets.append({"row": row_num, "title": title, "target_time": target_time_raw})

        wbs_lines = [
            line
            for line in clean_text(ws.cell(row_num, 8).value).split("\n")
            if line.strip()
        ]
        wbs_total += len(wbs_lines)

        description = clean_text(ws.cell(row_num, 3).value)
        goal = clean_text(ws.cell(row_num, 4).value)
        history = clean_text(ws.cell(row_num, 9).value)
        owner_raw = clean_text(ws.cell(row_num, 6).value)
        members_raw = clean_text(ws.cell(row_num, 7).value)

        notes = [
            {"id": "note_desc", "title": "說明", "content": description},
            {"id": "note_goal", "title": "達到目標", "content": goal},
        ]
        if history:
            notes.append({"id": "note_history", "title": "歷程紀錄", "content": history})

        tasks.append(
            {
                "row": row_num,
                "legacy": f"{IMPORT_PREFIX}-r{row_num:03d}",
                "title": title,
                "status_raw": status_raw,
                "status": status,
                "column": column,
                "description": description,
                "detail_notes": notes,
                "end_date": end_date,
                "target_time_raw": target_time_raw,
                "owner_raw": owner_raw,
                "members_raw": members_raw,
                "wbs_lines": [
                    {
                        "legacy": f"{IMPORT_PREFIX}-r{row_num:03d}-wbs{idx:03d}",
                        "source_line": line,
                        "depth": parse_wbs_line(line)[0],
                        "title": parse_wbs_line(line)[1],
                    }
                    for idx, line in enumerate(wbs_lines, start=1)
                ],
                "metadata": {
                    "source": SOURCE_LABEL,
                    "sourceSheet": SHEET_NAME,
                    "sourceRow": row_num,
                    "sourceLegacyPrefix": IMPORT_PREFIX,
                    "statusRaw": status_raw,
                    "targetTimeRaw": target_time_raw,
                    "assigneeRaw": owner_raw,
                    "membersRaw": members_raw,
                    "meetingControlMinutes": json_safe(ws.cell(row_num, 10).value),
                    "completionTimeRaw": json_safe(ws.cell(row_num, 11).value),
                },
            }
        )

    return {
        "source": str(path),
        "sheet": SHEET_NAME,
        "tasks": tasks,
        "summary": {
            "task_count": len(tasks),
            "wbs_count": wbs_total,
            "status_raw_counts": dict(status_raw_counts),
            "unknown_statuses": dict(unknown_statuses),
            "skipped_rows": skipped_rows,
            "non_date_targets": non_date_targets,
        },
    }


class SupabaseRest:
    def __init__(self, url: str, key: str) -> None:
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def request(
        self,
        method: str,
        table: str,
        params: dict[str, str] | None = None,
        body: Any | None = None,
        prefer: str | None = None,
    ) -> Any:
        query = urllib.parse.urlencode(params or {}, safe="(),.*")
        url = f"{self.base}/{table}"
        if query:
            url += f"?{query}"
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(url, data=data, method=method, headers=headers)
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                raw = response.read().decode("utf-8")
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="replace")
            raise ImportErrorWithContext(f"{method} {table} failed: HTTP {exc.code}: {details}") from exc

    def select_all(self, table: str, params: dict[str, str]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        start = 0
        page_size = 1000
        while True:
            page_params = dict(params)
            page_params["limit"] = str(page_size)
            page_params["offset"] = str(start)
            page = self.request("GET", table, page_params) or []
            rows.extend(page)
            if len(page) < page_size:
                return rows
            start += page_size

    def insert_one(self, table: str, row: dict[str, Any]) -> dict[str, Any]:
        result = self.request("POST", table, None, row, prefer="return=representation") or []
        if not result:
            raise ImportErrorWithContext(f"Insert into {table} returned no row.")
        return result[0]

    def delete_one(self, table: str, row_id: str) -> None:
        self.request("DELETE", table, {"id": f"eq.{row_id}"})


def in_filter(values: list[str]) -> str:
    return "in.(" + ",".join(values) + ")"


def resolve_target(
    db: SupabaseRest, email: str, board_name: str
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    profiles = db.select_all(
        "profiles",
        {"select": "id,email,display_name", "email": f"ilike.{email}"},
    )
    if len(profiles) != 1:
        raise ImportErrorWithContext(f"Expected one profile for {email}, found {len(profiles)}")
    profile = profiles[0]

    memberships = db.select_all(
        "tenant_members",
        {
            "select": "tenant_id,role,status,tenants(name)",
            "user_id": f"eq.{profile['id']}",
            "status": "eq.active",
        },
    )
    tenant_ids = sorted({row["tenant_id"] for row in memberships})
    if not tenant_ids:
        raise ImportErrorWithContext(f"Profile {email} has no active tenant memberships.")

    projects = db.select_all(
        "projects",
        {
            "select": "id,tenant_id,name,legacy_board_id,sort_order,created_at,updated_at",
            "tenant_id": in_filter(tenant_ids),
        },
    )
    exact = [project for project in projects if project.get("name") == board_name]
    if len(exact) == 1:
        return profile, exact[0], memberships
    if len(exact) > 1:
        raise ImportErrorWithContext(f"Board name {board_name!r} matched {len(exact)} projects.")

    fallback_names = {board_name.replace("看板", ""), DEFAULT_BOARD_FALLBACK}
    fallback = [project for project in projects if project.get("name") in fallback_names]
    if len(fallback) == 1:
        return profile, fallback[0], memberships
    if len(fallback) > 1:
        raise ImportErrorWithContext(f"Fallback board lookup matched {len(fallback)} projects.")

    raise ImportErrorWithContext(f"No accessible project matched {board_name!r}.")


def load_board_members(db: SupabaseRest, tenant_id: str, project_id: str) -> list[dict[str, Any]]:
    return db.select_all(
        "project_members",
        {
            "select": "project_id,tenant_id,user_id,role,profiles(id,email,display_name)",
            "tenant_id": f"eq.{tenant_id}",
            "project_id": f"eq.{project_id}",
        },
    )


def load_project_items(db: SupabaseRest, tenant_id: str, project_id: str) -> list[dict[str, Any]]:
    return db.select_all(
        "wbs_items",
        {
            "select": "*",
            "tenant_id": f"eq.{tenant_id}",
            "project_id": f"eq.{project_id}",
            "order": "sort_order.asc",
        },
    )


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=json_safe), encoding="utf-8")


def make_backup(
    db: SupabaseRest, out_dir: Path, profile: dict[str, Any], project: dict[str, Any]
) -> dict[str, Any]:
    tenant_id = project["tenant_id"]
    project_id = project["id"]
    backup = {
        "createdAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "profile": profile,
        "project": project,
        "wbs_items": load_project_items(db, tenant_id, project_id),
        "wbs_dependencies": db.select_all(
            "wbs_dependencies",
            {"select": "*", "tenant_id": f"eq.{tenant_id}", "project_id": f"eq.{project_id}"},
        ),
        "project_members": load_board_members(db, tenant_id, project_id),
        "board_role_permissions": db.select_all(
            "board_role_permissions",
            {"select": "*", "tenant_id": f"eq.{tenant_id}", "project_id": f"eq.{project_id}"},
        ),
    }
    write_json(out_dir / "backup.json", backup)
    return backup


def root_columns_by_title(items: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    roots = [item for item in items if not item.get("parent_id")]
    return {clean_text(item.get("title")): item for item in roots}


def import_rows(
    db: SupabaseRest,
    workbook: dict[str, Any],
    profile: dict[str, Any],
    project: dict[str, Any],
    members: list[dict[str, Any]],
    out_dir: Path,
    dry_run: bool,
) -> dict[str, Any]:
    tenant_id = project["tenant_id"]
    project_id = project["id"]
    member_index = member_match_index(members)
    current_items = load_project_items(db, tenant_id, project_id)
    by_legacy = {
        item["legacy_node_id"]: item
        for item in current_items
        if item.get("legacy_node_id")
    }
    columns = root_columns_by_title(current_items)

    used_columns = sorted({task["column"] for task in workbook["tasks"]})
    existing_root_max = max([int(item.get("sort_order") or 0) for item in current_items if not item.get("parent_id")] or [-1])
    planned_columns: dict[str, dict[str, Any]] = {}
    created_columns: list[str] = []

    for offset, title in enumerate(used_columns, start=1):
        if title in columns:
            planned_columns[title] = columns[title]
            continue
        definition = COLUMN_DEFS[title]
        existing = by_legacy.get(definition["legacy"])
        if existing:
            planned_columns[title] = existing
            continue
        row = {
            "tenant_id": tenant_id,
            "project_id": project_id,
            "legacy_node_id": definition["legacy"],
            "title": title,
            "description": None,
            "detail_notes": [],
            "status": definition["status"],
            "item_type": "group",
            "sort_order": existing_root_max + offset,
            "metadata": {
                "source": SOURCE_LABEL,
                "sourceLegacyPrefix": IMPORT_PREFIX,
                "createdForImport": True,
            },
            "created_by": profile["id"],
            "updated_by": profile["id"],
        }
        if dry_run:
            row["id"] = f"dry-run-{definition['legacy']}"
        else:
            row = db.insert_one("wbs_items", row)
            by_legacy[row["legacy_node_id"]] = row
            created_columns.append(title)
        planned_columns[title] = row

    children_by_parent: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in current_items:
        if item.get("parent_id"):
            children_by_parent[item["parent_id"]].append(item)

    inserted_cards = 0
    skipped_cards = 0
    inserted_wbs = 0
    skipped_wbs = 0
    unmatched_people: dict[str, set[str]] = {"owners": set(), "members": set()}
    card_status_counts: Counter[str] = Counter()

    for task in workbook["tasks"]:
        owner_ids, owner_unmatched = resolve_people(task["owner_raw"], member_index)
        member_ids, member_unmatched = resolve_people(task["members_raw"], member_index)
        unmatched_people["owners"].update(owner_unmatched)
        unmatched_people["members"].update(member_unmatched)

        column = planned_columns[task["column"]]
        column_id = column["id"]
        existing_siblings = children_by_parent.get(column_id, [])
        next_order = len(existing_siblings) + int(task["row"])
        card = by_legacy.get(task["legacy"])
        if card:
            skipped_cards += 1
        else:
            card_row = {
                "tenant_id": tenant_id,
                "project_id": project_id,
                "parent_id": None if dry_run else column_id,
                "legacy_node_id": task["legacy"],
                "title": task["title"],
                "description": task["description"] or None,
                "detail_notes": task["detail_notes"],
                "status": task["status"],
                "assignee_id": owner_ids[0] if owner_ids else None,
                "collaborator_ids": [uid for uid in member_ids if uid not in owner_ids],
                "end_date": task["end_date"],
                "item_type": "task",
                "sort_order": next_order,
                "metadata": task["metadata"],
                "created_by": profile["id"],
                "updated_by": profile["id"],
            }
            if dry_run:
                card = dict(card_row)
                card["id"] = f"dry-run-{task['legacy']}"
            else:
                card = db.insert_one("wbs_items", card_row)
                by_legacy[task["legacy"]] = card
                children_by_parent[column_id].append(card)
                inserted_cards += 1
        card_status_counts[task["status"]] += 1

        stack: dict[int, dict[str, Any]] = {0: card}
        order_by_parent: Counter[str] = Counter()
        for wbs in task["wbs_lines"]:
            existing_wbs = by_legacy.get(wbs["legacy"])
            if existing_wbs:
                skipped_wbs += 1
                stack[wbs["depth"]] = existing_wbs
                continue

            parent = stack.get(wbs["depth"] - 1) or card
            parent_id = parent["id"]
            order_by_parent[parent_id] += 1
            wbs_status = "completed" if task["status"] == "completed" else "todo"
            wbs_row = {
                "tenant_id": tenant_id,
                "project_id": project_id,
                "parent_id": None if dry_run else parent_id,
                "legacy_node_id": wbs["legacy"],
                "title": wbs["title"],
                "description": None,
                "detail_notes": [
                    {
                        "id": "note_source_line",
                        "title": "來源 WBS",
                        "content": wbs["source_line"],
                    }
                ],
                "status": wbs_status,
                "item_type": "task",
                "sort_order": order_by_parent[parent_id],
                "metadata": {
                    "source": SOURCE_LABEL,
                    "sourceSheet": SHEET_NAME,
                    "sourceRow": task["row"],
                    "sourceLegacyPrefix": IMPORT_PREFIX,
                    "sourceLine": wbs["source_line"],
                    "parsedDepth": wbs["depth"],
                },
                "created_by": profile["id"],
                "updated_by": profile["id"],
            }
            if dry_run:
                created = dict(wbs_row)
                created["id"] = f"dry-run-{wbs['legacy']}"
            else:
                created = db.insert_one("wbs_items", wbs_row)
                by_legacy[wbs["legacy"]] = created
                inserted_wbs += 1
            stack[wbs["depth"]] = created

    result = {
        "dryRun": dry_run,
        "target": {
            "email": profile["email"],
            "displayName": profile.get("display_name"),
            "tenantId": tenant_id,
            "projectId": project_id,
            "projectName": project["name"],
        },
        "source": workbook["summary"],
        "columns": {
            "used": used_columns,
            "created": created_columns,
        },
        "writeSummary": {
            "insertedCards": inserted_cards,
            "skippedExistingCards": skipped_cards,
            "insertedWbs": inserted_wbs,
            "skippedExistingWbs": skipped_wbs,
            "cardStatusCounts": dict(card_status_counts),
            "unmatchedPeople": {
                "owners": sorted(unmatched_people["owners"]),
                "members": sorted(unmatched_people["members"]),
            },
        },
    }
    write_json(out_dir / ("dry_run.json" if dry_run else "import_result.json"), result)
    return result


def validate_import(db: SupabaseRest, project: dict[str, Any], out_dir: Path) -> dict[str, Any]:
    items = load_project_items(db, project["tenant_id"], project["id"])
    imported = [
        item
        for item in items
        if clean_text(item.get("legacy_node_id")).startswith(IMPORT_PREFIX)
    ]
    legacy_counts = Counter(item.get("legacy_node_id") for item in imported)
    duplicates = sorted([legacy for legacy, count in legacy_counts.items() if count > 1])
    ids = {item["id"] for item in items}
    card_pattern = re.compile(rf"^{re.escape(IMPORT_PREFIX)}-r\d{{3}}$")
    wbs_pattern = re.compile(rf"^{re.escape(IMPORT_PREFIX)}-r\d{{3}}-wbs\d{{3}}$")
    cards = [item for item in imported if card_pattern.match(clean_text(item.get("legacy_node_id")))]
    wbs_items = [item for item in imported if wbs_pattern.match(clean_text(item.get("legacy_node_id")))]
    columns = [item for item in imported if "-column-" in clean_text(item.get("legacy_node_id"))]
    orphan_imports = [
        {
            "id": item["id"],
            "legacy_node_id": item.get("legacy_node_id"),
            "title": item.get("title"),
            "parent_id": item.get("parent_id"),
        }
        for item in imported
        if item.get("parent_id") and item.get("parent_id") not in ids
    ]
    missing_notes = [
        item.get("legacy_node_id")
        for item in cards
        if not {
            note.get("title")
            for note in (item.get("detail_notes") or [])
            if isinstance(note, dict)
        }.issuperset({"說明", "達到目標"})
    ]
    expected_status_by_raw = {"完成": "completed", "進行中": "in_progress"}
    status_mismatches = [
        {
            "legacy_node_id": item.get("legacy_node_id"),
            "title": item.get("title"),
            "statusRaw": (item.get("metadata") or {}).get("statusRaw"),
            "expected": expected_status_by_raw.get((item.get("metadata") or {}).get("statusRaw")),
            "actual": item.get("status"),
            "end_date": item.get("end_date"),
        }
        for item in cards
        if expected_status_by_raw.get((item.get("metadata") or {}).get("statusRaw"))
        and expected_status_by_raw[(item.get("metadata") or {}).get("statusRaw")] != item.get("status")
    ]
    validation = {
        "importPrefix": IMPORT_PREFIX,
        "totalPrefixItems": len(imported),
        "columns": len(columns),
        "cards": len(cards),
        "wbsItems": len(wbs_items),
        "statusCounts": dict(Counter(item.get("status") for item in cards)),
        "statusMismatches": status_mismatches,
        "duplicates": duplicates,
        "orphanImports": orphan_imports,
        "cardsMissingRequiredNotes": missing_notes,
        "passed": (
            len(cards) == 20
            and len(wbs_items) == 178
            and not duplicates
            and not orphan_imports
            and not missing_notes
        ),
    }
    write_json(out_dir / "validation.json", validation)
    return validation


def rollback_import(db: SupabaseRest, project: dict[str, Any], out_dir: Path, yes: bool) -> dict[str, Any]:
    if not yes:
        raise ImportErrorWithContext("Rollback requires --yes.")
    items = load_project_items(db, project["tenant_id"], project["id"])
    imported = [
        item
        for item in items
        if clean_text(item.get("legacy_node_id")).startswith(IMPORT_PREFIX)
    ]
    children: dict[str | None, list[dict[str, Any]]] = defaultdict(list)
    by_id = {item["id"]: item for item in imported}
    for item in imported:
        children[item.get("parent_id")].append(item)

    def depth(item: dict[str, Any]) -> int:
        value = 0
        current = item
        seen = set()
        while current.get("parent_id") in by_id and current.get("parent_id") not in seen:
            seen.add(current["parent_id"])
            current = by_id[current["parent_id"]]
            value += 1
        return value

    deleted: list[dict[str, Any]] = []
    for item in sorted(imported, key=depth, reverse=True):
        db.delete_one("wbs_items", item["id"])
        deleted.append({"id": item["id"], "legacy_node_id": item.get("legacy_node_id"), "title": item.get("title")})
    result = {"deleted": deleted, "count": len(deleted)}
    write_json(out_dir / "rollback_result.json", result)
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--env", default=DEFAULT_ENV)
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--email", default=DEFAULT_EMAIL)
    parser.add_argument("--board", default=DEFAULT_BOARD_NAME)
    parser.add_argument("--out-dir", default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--rollback", action="store_true")
    parser.add_argument("--yes", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    env_path = Path(args.env)
    workbook_path = Path(args.xlsx)
    if not workbook_path.exists():
        matches = list(Path("..").glob("Meeting Minutes_*_2026Q2.xlsx"))
        if len(matches) == 1:
            workbook_path = matches[0]
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.out_dir) if args.out_dir else Path("outputs") / "meeting_minutes_import_2026q2" / timestamp

    env = load_env(env_path)
    supabase_url = env.get("SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise ImportErrorWithContext(f"{env_path} must contain SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY.")
    db = SupabaseRest(supabase_url, service_key)

    profile, project, memberships = resolve_target(db, args.email, args.board)
    out_dir.mkdir(parents=True, exist_ok=True)
    write_json(out_dir / "target.json", {"profile": profile, "project": project, "memberships": memberships})

    if args.rollback:
        result = rollback_import(db, project, out_dir, args.yes)
        print(json.dumps({"mode": "rollback", "outDir": str(out_dir), **result}, ensure_ascii=False, indent=2))
        return 0

    workbook = parse_workbook(workbook_path)
    members = load_board_members(db, project["tenant_id"], project["id"])
    write_json(out_dir / "parsed_workbook.json", workbook)
    write_json(
        out_dir / "board_members.json",
        [
            {
                "user_id": member.get("user_id"),
                "role": member.get("role"),
                "email": (member.get("profiles") or {}).get("email"),
                "display_name": (member.get("profiles") or {}).get("display_name"),
            }
            for member in members
        ],
    )

    if not args.dry_run:
        make_backup(db, out_dir, profile, project)

    result = import_rows(db, workbook, profile, project, members, out_dir, args.dry_run)
    validation = None if args.dry_run else validate_import(db, project, out_dir)
    print(json.dumps({"mode": "dry-run" if args.dry_run else "import", "outDir": str(out_dir), "result": result, "validation": validation}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ImportErrorWithContext as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
